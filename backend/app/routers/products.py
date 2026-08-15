from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.models.product import Product
from app.models.business import Business
from app.schemas.product import ProductCreate, ProductOut, ProductUpdate
from app.core.deps import get_current_business

router = APIRouter(prefix="/products", tags=["products"])

LOW_STOCK_THRESHOLD = 5

def check_low_stock(stock_qty: int) -> bool:
    return stock_qty < LOW_STOCK_THRESHOLD

@router.get("/import-template")
def download_import_template():
    """Generate and return an Excel template for product import"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Import Template"
    
    headers = ['name', 'category', 'price', 'stock_qty', 'unit']
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    examples = [
        {'name': 'Panadol', 'category': 'Medicine', 'price': 50.0, 'stock_qty': 100, 'unit': 'pcs'},
        {'name': 'Cough Syrup', 'category': 'Medicine', 'price': 150.0, 'stock_qty': 50, 'unit': 'bottle'},
        {'name': 'Bandage', 'category': 'Medical Supplies', 'price': 20.0, 'stock_qty': 200, 'unit': 'pcs'}
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=example.get(header, ''))
    
    ws2 = wb.create_sheet("Instructions")
    
    instructions = [
        "Product Import Template - Instructions",
        "",
        "Required Columns (must be present):",
        "1. name - Product name (required)",
        "2. price - Product price in numbers (required)",
        "",
        "Optional Columns:",
        "3. category - Product category (optional)",
        "4. stock_qty - Available stock quantity (optional, defaults to 0)",
        "5. unit - Unit of measurement e.g., pcs, kg, box (optional)",
        "",
        "Rules:",
        "- Column names are case-insensitive",
        "- Price must be greater than 0",
        "- Stock quantity cannot be negative",
        "- Duplicate product names will be skipped",
        "- Delete example rows before importing your own data",
        "",
        "Example:",
        "| name | category | price | stock_qty | unit |",
        "| Panadol | Medicine | 50 | 100 | pcs |"
    ]
    
    for row_idx, text in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws2.column_dimensions['A'].width = 60
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=product_import_template.xlsx"
        }
    )

@router.post("/import", status_code=status.HTTP_200_OK)
async def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_business: Business = Depends(get_current_business)
):
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload .csv, .xlsx, or .xls file")
    
    try:
        content = await file.read()
        
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        df.columns = [str(col).lower().strip() for col in df.columns]
        
        required_columns = ['name', 'price']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_columns)}")
        
        imported = 0
        failed = 0
        skipped_duplicates = 0
        errors = []
        
        existing_products = db.query(Product).filter(Product.business_id == current_business.id).all()
        existing_names = {p.name.lower(): p for p in existing_products}
        
        for index, row in df.iterrows():
            row_num = index + 2
            try:
                name = str(row.get('name', '')).strip()
                if not name or name == 'nan':
                    errors.append({"row": row_num, "error": "Name is required"})
                    failed += 1
                    continue
                
                price_raw = row.get('price')
                try:
                    price = float(price_raw)
                    if price <= 0:
                        errors.append({"row": row_num, "error": f"Price must be greater than 0"})
                        failed += 1
                        continue
                except (ValueError, TypeError):
                    errors.append({"row": row_num, "error": f"Invalid price value: '{price_raw}'"})
                    failed += 1
                    continue
                
                stock_raw = row.get('stock_qty', 0)
                try:
                    if pd.isna(stock_raw) or str(stock_raw).strip() == '':
                        stock_qty = 0
                    else:
                        stock_qty = int(float(stock_raw))
                        if stock_qty < 0:
                            errors.append({"row": row_num, "error": "Stock cannot be negative"})
                            failed += 1
                            continue
                except (ValueError, TypeError):
                    errors.append({"row": row_num, "error": f"Invalid stock quantity: '{stock_raw}'"})
                    failed += 1
                    continue
                
                category_raw = row.get('category')
                category = str(category_raw).strip() if pd.notna(category_raw) and str(category_raw).strip() != 'nan' else None
                
                unit_raw = row.get('unit')
                unit = str(unit_raw).strip() if pd.notna(unit_raw) and str(unit_raw).strip() != 'nan' else None
                
                name_lower = name.lower()
                if name_lower in existing_names:
                    skipped_duplicates += 1
                    errors.append({"row": row_num, "error": f"Duplicate product name '{name}' — skipped"})
                    continue
                
                db_product = Product(
                    business_id=current_business.id,
                    name=name,
                    category=category,
                    price=price,
                    stock_qty=stock_qty,
                    unit=unit
                )
                
                db.add(db_product)
                existing_names[name_lower] = db_product
                imported += 1
                
            except Exception as e:
                errors.append({"row": row_num, "error": f"Unexpected error: {str(e)}"})
                failed += 1
        
        if imported > 0:
            db.commit()
        else:
            db.rollback()
        
        return {
            "total_rows": len(df),
            "imported": imported,
            "failed": failed,
            "skipped_duplicates": skipped_duplicates,
            "errors": errors[:50],
            "message": f"Import complete: {imported} imported, {failed} failed, {skipped_duplicates} duplicates skipped"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to import products: {str(e)}")

@router.post("/", response_model=ProductOut, status_code=status.HTTP_201_CREATED)
def create_product(
    product_data: ProductCreate,
    db: Session = Depends(get_db),
    current_business: Business = Depends(get_current_business)
):
    try:
        db_product = Product(
            business_id=current_business.id,
            name=product_data.name,
            category=product_data.category,
            price=product_data.price,
            stock_qty=product_data.stock_qty,
            unit=product_data.unit,
            attributes=product_data.attributes
        )
        
        db.add(db_product)
        db.commit()
        db.refresh(db_product)
        
        product_out = ProductOut.from_orm(db_product)
        product_out.low_stock = check_low_stock(db_product.stock_qty)
        
        return product_out
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create product: {str(e)}")

@router.get("/", response_model=List[ProductOut])
def list_products(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    category: Optional[str] = Query(None),
    low_stock_only: bool = Query(False),
    db: Session = Depends(get_db),
    current_business: Business = Depends(get_current_business)
):
    query = db.query(Product).filter(Product.business_id == current_business.id)
    
    if search:
        search_term = f"%{search}%"
        query = query.filter(or_(Product.name.ilike(search_term), Product.category.ilike(search_term)))
    
    if category:
        query = query.filter(Product.category == category)
    
    if low_stock_only:
        query = query.filter(Product.stock_qty < LOW_STOCK_THRESHOLD)
    
    offset = (page - 1) * limit
    products = query.order_by(Product.created_at.desc()).offset(offset).limit(limit).all()
    
    result = []
    for product in products:
        product_out = ProductOut.from_orm(product)
        product_out.low_stock = check_low_stock(product.stock_qty)
        result.append(product_out)
    
    return result

@router.get("/{product_id}", response_model=ProductOut)
def get_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_business: Business = Depends(get_current_business)
):
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.business_id == current_business.id
    ).first()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product_out = ProductOut.from_orm(product)
    product_out.low_stock = check_low_stock(product.stock_qty)
    
    return product_out

@router.put("/{product_id}", response_model=ProductOut)
def update_product(
    product_id: int,
    product_data: ProductUpdate,
    db: Session = Depends(get_db),
    current_business: Business = Depends(get_current_business)
):
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.business_id == current_business.id
    ).first()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    try:
        update_data = product_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(product, field, value)
        
        db.commit()
        db.refresh(product)
        
        product_out = ProductOut.from_orm(product)
        product_out.low_stock = check_low_stock(product.stock_qty)
        
        return product_out
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update product: {str(e)}")

@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_business: Business = Depends(get_current_business)
):
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.business_id == current_business.id
    ).first()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    try:
        db.delete(product)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete product: {str(e)}")